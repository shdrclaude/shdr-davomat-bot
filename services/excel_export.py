"""Oylik davomatni Excel (.xlsx) ga eksport qilish."""
from __future__ import annotations

import calendar as _cal
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Attendance, Employee, EmployeeStatus
from utils.time_helpers import MONTHS_UZ, fmt_time

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LATE_FILL = PatternFill("solid", fgColor="FCE4D6")


async def build_month_excel(session: AsyncSession, branch_id: int, year: int, month: int) -> bytes:
    first = date(year, month, 1)
    days_in_month = _cal.monthrange(year, month)[1]
    last = date(year, month, days_in_month)

    res = await session.execute(
        select(Employee).where(
            and_(Employee.branch_id == branch_id, Employee.status == EmployeeStatus.faol)
        ).order_by(Employee.full_name)
    )
    employees = list(res.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTHS_UZ[month].capitalize()} {year}"

    headers = ["Xodim", "Lavozim"] + [str(d) for d in range(1, days_in_month + 1)] + [
        "Ishlagan kun", "Kech (marta)", "Jami soat"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for emp in employees:
        ares = await session.execute(
            select(Attendance).where(
                and_(
                    Attendance.employee_id == emp.id,
                    Attendance.date >= first,
                    Attendance.date <= last,
                )
            )
        )
        by_day = {a.date.day: a for a in ares.scalars().all()}
        row = [emp.full_name, emp.position or "—"]
        worked_days = late_count = total_minutes = 0
        for d in range(1, days_in_month + 1):
            a = by_day.get(d)
            if a and a.check_in:
                worked_days += 1
                total_minutes += a.worked_minutes or 0
                if a.is_late:
                    late_count += 1
                row.append(fmt_time(a.check_in))
            else:
                row.append("")
        row += [worked_days, late_count, round(total_minutes / 60, 1)]
        ws.append(row)
        # kech qolgan kunlarni bo'yash
        r = ws.max_row
        for idx, d in enumerate(range(1, days_in_month + 1)):
            a = by_day.get(d)
            if a and a.is_late:
                ws.cell(row=r, column=3 + idx).fill = LATE_FILL

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
